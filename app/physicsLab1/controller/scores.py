import os
from pathlib import Path

import excel2img

import xlsxwriter
from telegram import Update
from telegram.ext import CallbackContext

from app.physicsLab1.api import get_work_by_user_id, get_user, get_work_list_all, get_work_all, get_work_by_id
from app.physicsLab1.main import set_user_to_user_data, reset
from app.user.api import get_user as get_user_db

import openpyxl

from core.style.InlineKeyboardMarkup import get_ikm_physics_lab_1_user_list, \
    get_ikm_physics_lab_1_work_list_for_set_score

parent_path = Path(__file__).resolve().parent


# function to remove empty rows

def remove(sheet, row):
    for cell in row:
        if cell.value is not None:
            return
    sheet.delete_rows(row[0].row, 1)


def get_one_person_score_by_user(update: Update, context: CallbackContext):
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=update.effective_message.message_id)

    get_one_person_score_by_id(update, context, update.effective_user.id)


def get_one_person_score_by_id(update: Update, context: CallbackContext, id_in: int):
    user = get_user(id_in=id_in)

    update.effective_message.reply_chat_action('upload_photo')
    message = update.effective_message.reply_text("uploading photo")
    path = '../junk/'

    if not os.path.exists(parent_path.joinpath(path)):
        os.makedirs(parent_path.joinpath(path))

    path = parent_path.joinpath(path).resolve()
    path_xlsx = str(
        path.joinpath('./physicsLab1_' + user.user_rel.full_name + '_' + str(user.student_number) + '.xlsx'))
    path_png = str(path.joinpath('./physicsLab1_' + user.user_rel.full_name + '_' + str(user.student_number) + '.png'))

    works = get_work_by_user_id(user.user_id)

    workbook = xlsxwriter.Workbook(path_xlsx)

    worksheet = workbook.add_worksheet('scores')
    cell_format = workbook.add_format()

    cell_format.set_align('center')
    cell_format.set_align('vcenter')

    index = 0

    worksheet.write(index, 0, user.user_rel.full_name, cell_format)
    worksheet.write(index, 1, 'نام', cell_format)
    index += 1

    worksheet.write(index, 0, get_user(id_in=user.user_rel.id).student_number, cell_format)
    worksheet.write(index, 1, 'شماره دانشجویی', cell_format)
    index += 1

    worksheet.write(index, 1, 'نام آزمایش', cell_format)
    worksheet.write(index, 0, 'نمره', cell_format)
    index += 1

    for work in works:
        worksheet.write(index, 1, work.work.work_name, cell_format)
        if work.score is None:
            worksheet.write(index, 0, 'تصحیح نشده', cell_format)
        else:
            worksheet.write(index, 0, work.score, cell_format)
        index += 1

    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:B', 40)
    workbook.close()
    excel2img.export_img(path_xlsx, path_png)

    try:
        update.effective_message.reply_document(open(path_xlsx, 'rb'))
        update.effective_message.reply_photo(open(path_png, 'rb'))
    except Exception as e:
        print(e)
        update.effective_message.reply_text("something went wrong,contact admin")

    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=message.message_id)
    os.remove(path_png)
    os.remove(path_xlsx)


def get_all_score(update: Update, context: CallbackContext):
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=update.effective_message.message_id)
    update.effective_message.reply_chat_action('upload_document')
    message = update.effective_message.reply_text("uploading document")
    path = '../junk/'

    if not os.path.exists(parent_path.joinpath(path)):
        os.makedirs(parent_path.joinpath(path))

    path = parent_path.joinpath(path).resolve()

    path_xlsx = str(path.joinpath('./physicsLab1_all.xlsx'))

    # path_png = str(path.joinpath('./physicsLab1_all.png'))

    works_list = get_work_list_all()

    workbook = xlsxwriter.Workbook(path_xlsx)

    worksheet = workbook.add_worksheet('scores')
    cell_format = workbook.add_format()

    cell_format.set_align('center')
    cell_format.set_align('vcenter')

    index = 0

    worksheet.write(0, index, 'نام', cell_format)
    index += 1

    worksheet.write(0, index, 'شماره دانشجویی', cell_format)
    index += 1

    for work in works_list:
        worksheet.write(0, index, work.work_name, cell_format)
        index += 1

    works = get_work_all()

    for work in works:
        worksheet.write(work.user_id, 0, work.user_rel.full_name, cell_format)
        worksheet.write(work.user_id, 1, get_user(id_in=work.user_rel.id).student_number, cell_format)
        if work.score is None:
            worksheet.write(work.user_id, work.work.id + 1, 'تصحیح نشده', cell_format)
        else:
            worksheet.write(work.user_id, work.work.id + 1, work.score, cell_format)

    worksheet.set_column('A:W', 40)
    workbook.close()

    wb = openpyxl.load_workbook(path_xlsx)
    sheet = wb['scores']
    for i in range(20):
        for row in sheet:
            remove(sheet, row)
    wb.save(path_xlsx)
    # excel2img.export_img(path_xlsx, path_png)

    try:
        update.effective_message.reply_document(open(path_xlsx, 'rb'))
        # update.effective_message.reply_photo(open(path_png, 'rb'))
    except Exception as e:
        print(e)
        update.effective_message.reply_text("something went wrong,contact admin")

    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=message.message_id)

    os.remove(path_xlsx)
    # os.remove(path_png)


def get_score(update: Update, context: CallbackContext):
    user = get_user(user=update.effective_user)
    user_db = get_user_db(user=update.effective_message.from_user)

    if (user.user_id == 0 or user.id is None) and not user_db.check_admin():
        update.effective_message.reply_text("your are not an engy user,please start bot with command /start to join")
        return

    if not user_db.check_admin():
        get_one_person_score_by_user(update, context)
    else:
        get_all_score(update, context)


def set_score(update: Update, context: CallbackContext):
    user_db = get_user_db(user=update.effective_message.from_user)

    set_user_to_user_data(update, context)
    user_data = context.user_data

    if not user_db.check_admin():
        update.effective_message.reply_text("your are not allow to use this")
        return

    chat_data = context.chat_data
    chat_data['app'] = 'physics_lab_1'

    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=update.effective_message.message_id)

    if user_data['is_admin_flag']:
        reply_markup = get_ikm_physics_lab_1_user_list()
        message = update.message.reply_text("select student who want to set his score",
                                            reply_markup=reply_markup)

        chat_data['physics_lab_1_message_id'] = message.message_id
        chat_data['command'] = 'select_user_for_set_score'
        return


def select_user_for_set_score(update: Update, context: CallbackContext):
    query = update.callback_query
    chat_data = context.chat_data
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=chat_data['physics_lab_1_message_id'])
    query.answer()

    user = get_user_db(user_id=int(query.data))

    reply_markup = get_ikm_physics_lab_1_work_list_for_set_score(user.user_id)
    if len(reply_markup.inline_keyboard):
        message = update.effective_message.reply_text("please select a HW ,whose you want to set score",
                                                      reply_markup=reply_markup)
        chat_data['command'] = 'select_work_set_score'
        chat_data['data'] = {'user_id': int(query.data)}
        chat_data['physics_lab_1_message_id'] = message.message_id
    else:
        update.effective_message.reply_text("student upload nothing")
        reset(update, context)

    reset(update, context)


def get_score_for_set_score(update: Update, context: CallbackContext):
    query = update.callback_query
    chat_data = context.chat_data
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=chat_data['physics_lab_1_message_id'])
    chat_data['data']['work_id'] = int(query.data)

    work = get_work_by_id(chat_data['data']['work_id'])
    message = update.effective_message.reply_text(
        "You have been selected work " + work.work.work_name + "\nplease enter your score")

    chat_data['command'] = 'enter_work_score'
    chat_data['physics_lab_1_message_id'] = message.message_id
    query.answer()


def enter_score(update: Update, context: CallbackContext):
    chat_data = context.chat_data
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=chat_data['physics_lab_1_message_id'])
    context.bot.delete_message(chat_id=update.effective_chat.id, message_id=update.effective_message.message_id)

    work = get_work_by_id(chat_data['data']['work_id'])
    try:
        work.update_score(
            int(update.effective_message.text))  # TODO:bayad age int vared nakarde bod dobare biad to hamin marhale v bg int vared kon
    except:
        update.effective_message.reply_text("something went wrong")
    user = get_user(user_id=work.user_id)
    update.effective_message.reply_text(
        "You have been updated score of work " + work.work.work_name + " from " + user.full_name + " to " + str(
            update.effective_message.text))
    reset(update, context)
